from datetime import date
import io
import re
from fpdf import FPDF
from groq import Groq
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import plotly.express as px
import streamlit as st
from streamlit_gsheets import GSheetsConnection

# Set Halaman Streamlit
st.set_page_config(
    page_title="Employee Database Manager", page_icon="👥", layout="wide"
)

# --- SEMBUNYIKAN MENU KANAN ATAS ---
st.markdown(
    """
    <style>
    div[data-testid="stToolbarActions"] { display: none !important; }
    div[data-testid="stDecoration"] { display: none !important; }
    #MainMenu { visibility: hidden !important; }
    [data-testid="stCollapsedControl"] { display: flex !important; visibility: visible !important; }
    </style>
    """,
    unsafe_allow_html=True,
)


# --- FUNGSI FORMAT ANGKA GRAFIK (RUPIAH & SINGKATAN LOKAL) ---
def format_rp_short(val):
    """Mengubah nominal angka besar menjadi format singkat Rp Indonesia (M / Juta)."""
    if val >= 1e9:
        return f"Rp {val/1e9:.2f} M".replace(".", ",")
    elif val >= 1e6:
        return f"Rp {val/1e6:.1f} Jt".replace(".", ",")
    elif val > 0:
        return f"Rp {val:,.0f}".replace(",", ".")
    return "Rp 0"


# --- SISTEM PROTEKSI PASSWORD APPS ---
def check_password():
    """Memeriksa apakah pengguna sudah memasukkan password aplikasi yang benar."""

    def password_entered():
        if st.session_state.get("app_password_input") == st.secrets.get(
            "PASSWORD"
        ):
            st.session_state["app_password_correct"] = True
            if "app_password_input" in st.session_state:
                del st.session_state["app_password_input"]
        else:
            st.session_state["app_password_correct"] = False

    if not st.session_state.get("app_password_correct", False):
        st.markdown("<br>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.title("🔒 Employee Database Manager")
            st.caption("Silakan masukkan password untuk mengakses aplikasi.")

            st.text_input(
                "Masukkan Password:",
                type="password",
                on_change=password_entered,
                key="app_password_input",
            )

            if (
                "app_password_correct" in st.session_state
                and not st.session_state["app_password_correct"]
            ):
                st.error("❌ Password salah. Silakan coba lagi.")

        return False
    return True


if not check_password():
    st.stop()


# --- FUNGSI PROTEKSI PASSWORD MANPOWER COST ---
def check_manpower_access():
    """Memeriksa otorisasi password khusus sebelum memberikan akses ke Manpower Cost Manager."""
    if "manpower_authenticated" not in st.session_state:
        st.session_state["manpower_authenticated"] = False

    if st.session_state["manpower_authenticated"]:
        return True

    st.title("🔒 Manpower Cost Manager")
    st.warning(
        "Halaman ini berisi data sensitif finansial dan rahasia perusahaan."
    )

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        with st.form("form_manpower_auth", clear_on_submit=True):
            st.subheader("Otorisasi Diperlukan")
            input_pass = st.text_input(
                "Masukkan Password Akses Manpower Cost:", type="password"
            )
            btn_submit = st.form_submit_button(
                "Buka Akses Dashboard", use_container_width=True
            )

            if btn_submit:
                correct_pass = st.secrets.get(
                    "MANPOWER_PASSWORD", "PasswordManpower2026!"
                )

                if input_pass == correct_pass:
                    st.session_state["manpower_authenticated"] = True
                    st.success("Akses Diterima! Memuat data...")
                    st.rerun()
                else:
                    st.error("❌ Password salah. Silakan coba lagi.")

    return False


# ==============================================================================
# NAVIGASI & AKSES UTAMA
# ==============================================================================

if st.sidebar.button("🚪 Keluar Aplikasi"):
    st.session_state["app_password_correct"] = False
    st.session_state["manpower_authenticated"] = False
    st.rerun()

st.sidebar.markdown("---")

# --- PIN ADMINISTRATOR ---
ADMIN_PIN = st.secrets.get("ADMIN_PIN", "2273")

# --- KONEKSI GOOGLE SHEETS ---
conn = st.connection("gsheets", type=GSheetsConnection)

# --- SIDEBAR: NAVIGASI MODUL ---
st.sidebar.header("📁 Menu Utama")
menu_pilihan = st.sidebar.radio(
    "Pilih Halaman:",
    [
        "👥 Master Data Karyawan",
        "⏱️ Rekap Absensi (Timesheet)",
        "💳 Manpower Cost Manager",
        "🤖 AI HR Assistant",
    ],
)
st.sidebar.markdown("---")


# ==============================================================================
# FUNGSI HELPER MASTER KARYAWAN
# ==============================================================================


def load_data():
    try:
        df = conn.read(worksheet="Master_Karyawan", ttl=0)
        if df is not None and not df.empty:
            if "ID" in df.columns:
                df["ID"] = df["ID"].astype(str).str.strip().str.upper()
            if "Jabatan" in df.columns and "Posisi" not in df.columns:
                df.rename(columns={"Jabatan": "Posisi"}, inplace=True)
            if "Site" not in df.columns:
                df["Site"] = ""
            if "Status" not in df.columns:
                df["Status"] = "Aktif"
            if "Tanggal Resign" not in df.columns:
                df["Tanggal Resign"] = "-"
            if "Terakhir Diperbarui" not in df.columns:
                df["Terakhir Diperbarui"] = str(date.today())
        return df if df is not None else pd.DataFrame()
    except Exception:
        try:
            df = conn.read(ttl=0)
            if df is not None and not df.empty:
                if "ID" in df.columns:
                    df["ID"] = df["ID"].astype(str).str.strip().str.upper()
                if "Jabatan" in df.columns and "Posisi" not in df.columns:
                    df.rename(columns={"Jabatan": "Posisi"}, inplace=True)
                if "Site" not in df.columns:
                    df["Site"] = ""
                if "Status" not in df.columns:
                    df["Status"] = "Aktif"
                if "Tanggal Resign" not in df.columns:
                    df["Tanggal Resign"] = "-"
                if "Terakhir Diperbarui" not in df.columns:
                    df["Terakhir Diperbarui"] = str(date.today())
            return df if df is not None else pd.DataFrame()
        except Exception as e:
            st.error(f"Gagal terhubung ke Google Sheets: {e}")
            return pd.DataFrame(
                columns=[
                    "ID",
                    "Nama Lengkap",
                    "Posisi",
                    "Cost Center",
                    "Tanggal Bergabung",
                    "Akhir Kontrak",
                    "Tanggal Resign",
                    "Site",
                    "Status",
                    "Terakhir Diperbarui",
                ]
            )


def load_snapshot_data():
    try:
        df_snap = conn.read(worksheet="Snapshot_Bulanan", ttl=0)
        return df_snap if df_snap is not None else pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def save_data(df):
    clean_df = df.fillna("")
    conn.update(worksheet="Master_Karyawan", data=clean_df)
    st.session_state.employees = clean_df


def generate_pdf(df):
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)

    pdf.cell(
        0, 10, "LAPORAN DATABASE KARYAWAN", new_x="LMARGIN", new_y="NEXT", align="C"
    )
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(
        0,
        6,
        f"Dicetak Tanggal: {date.today().strftime('%d-%m-%Y')} | Total Record:"
        f" {len(df)}",
        new_x="LMARGIN",
        new_y="NEXT",
        align="C",
    )
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(230, 230, 230)

    col_widths = [22, 45, 35, 25, 25, 25, 25, 25, 20, 30]
    headers = [
        "ID",
        "Nama Lengkap",
        "Posisi",
        "Cost Center",
        "Tgl Join",
        "End Kontrak",
        "Tgl Resign",
        "Site",
        "Status",
        "Updated",
    ]

    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 7, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for _, row in df.iterrows():
        pdf.cell(col_widths[0], 6, str(row.get("ID", "")), border=1, align="C")
        pdf.cell(
            col_widths[1], 6, str(row.get("Nama Lengkap", ""))[:25], border=1
        )
        pdf.cell(col_widths[2], 6, str(row.get("Posisi", ""))[:20], border=1)
        pdf.cell(
            col_widths[3],
            6,
            str(row.get("Cost Center", "")),
            border=1,
            align="C",
        )
        pdf.cell(
            col_widths[4],
            6,
            str(row.get("Tanggal Bergabung", "")),
            border=1,
            align="C",
        )
        pdf.cell(
            col_widths[5],
            6,
            str(row.get("Akhir Kontrak", "")),
            border=1,
            align="C",
        )
        pdf.cell(
            col_widths[6],
            6,
            str(row.get("Tanggal Resign", "-")),
            border=1,
            align="C",
        )
        pdf.cell(col_widths[7], 6, str(row.get("Site", "")), border=1, align="C")
        pdf.cell(
            col_widths[8],
            6,
            str(row.get("Status", "Aktif")),
            border=1,
            align="C",
        )
        pdf.cell(
            col_widths[9],
            6,
            str(row.get("Terakhir Diperbarui", "")),
            border=1,
            align="C",
        )
        pdf.ln()

    out = pdf.output()
    return bytes(out) if isinstance(out, (str, bytearray)) else out


def generate_excel_formatted(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Karyawan"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    data_font = Font(name="Calibri", size=10)
    border_thin = Side(border_style="thin", color="D9D9D9")
    border_box = Border(
        left=border_thin,
        right=border_thin,
        top=border_thin,
        bottom=border_thin,
    )

    ws.merge_cells("A1:J1")
    ws["A1"] = "LAPORAN DATABASE KARYAWAN"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = (
        f"Tanggal Ekspor: {date.today().strftime('%d-%m-%Y')} | Total Record:"
        f" {len(df)}"
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18

    headers = list(df.columns)
    ws.append([])
    ws.append(headers)
    ws.row_dimensions[4].height = 24

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for _, row in df.iterrows():
        row_data = list(row)
        ws.append(row_data)
        row_num = ws.max_row
        ws.row_dimensions[row_num].height = 20

        for c_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=c_idx)
            cell.font = data_font
            cell.border = border_box
            if headers[c_idx - 1] in [
                "ID",
                "Tanggal Bergabung",
                "Akhir Kontrak",
                "Tanggal Resign",
                "Status",
                "Terakhir Diperbarui",
            ]:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row >= 4 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


if "employees" not in st.session_state:
    st.session_state.employees = load_data()


def generate_next_id():
    df = st.session_state.employees
    max_num = 0
    if not df.empty and "ID" in df.columns:
        for emp_id in df["ID"]:
            if str(emp_id).startswith("EMP-"):
                try:
                    num = int(str(emp_id).split("-")[1])
                    if num > max_num:
                        max_num = num
                except ValueError:
                    pass
    return f"EMP-{str(max_num + 1).zfill(3)}"


# --- SIDEBAR: OTENTIKASI ADMIN ---
st.sidebar.header("🔐 Akses Pengguna")
role = st.sidebar.radio(
    "Pilih Mode Akses:", ["Umum (View Only)", "Administrator"]
)

is_admin = False
if role == "Administrator":
    pin_input = st.sidebar.text_input("Masukkan PIN Admin:", type="password")
    if pin_input == ADMIN_PIN:
        st.sidebar.success("Akses Administrator Aktif!")
        is_admin = True
    elif pin_input != "":
        st.sidebar.error("PIN Salah!")
    else:
        st.sidebar.info("Masukkan PIN Administrator.")


# ==============================================================================
# MODUL 1: MASTER DATA KARYAWAN
# ==============================================================================
if menu_pilihan == "👥 Master Data Karyawan":

    st.title("Employee Database Manager")
    st.caption("Created by iqbalmantam")

    df_master_current = st.session_state.employees
    total_karyawan = len(df_master_current)
    total_aktif = (
        len(df_master_current[df_master_current["Status"] == "Aktif"])
        if "Status" in df_master_current.columns
        else total_karyawan
    )
    total_resign = (
        len(df_master_current[df_master_current["Status"] == "Resign"])
        if "Status" in df_master_current.columns
        else 0
    )

    col_m1, col_m2, col_m3 = st.columns(3)
    with col_m1:
        st.metric(label="Karyawan Aktif", value=total_aktif)
    with col_m2:
        st.metric(label="Karyawan Resign", value=total_resign)
    with col_m3:
        st.metric(label="Total Record Data", value=total_karyawan)

    st.divider()

    if is_admin:
        st.sidebar.markdown("---")
        st.sidebar.header("⚡ Kontrol Admin (Master)")

        if st.sidebar.button("🔄 Sync / Refresh Data Master"):
            st.session_state.employees = load_data()
            st.rerun()

        # 1. Tambah Karyawan Baru
        with st.sidebar.expander("➕ Tambah Karyawan Baru", expanded=False):
            with st.form("add_employee_form", clear_on_submit=True):
                auto_id = generate_next_id()
                new_id = st.text_input("ID Karyawan", value=auto_id)
                new_name = st.text_input("Nama Lengkap")
                new_role = st.text_input("Posisi")
                new_cc = st.text_input("Cost Center", placeholder="CC-101")
                new_join = st.date_input(
                    "Tanggal Bergabung", value=date.today()
                )
                new_end = st.date_input("Akhir Kontrak", value=date.today())
                new_site = st.text_input(
                    "Site / Lokasi Kerja",
                    placeholder="Contoh: JDC / Head Office",
                )
                new_status = st.selectbox(
                    "Status Karyawan", ["Aktif", "Resign", "PKWT"]
                )

                new_resign_date = "-"
                if new_status == "Resign":
                    new_resign_date = st.date_input(
                        "Tanggal Resign", value=date.today()
                    ).strftime("%Y-%m-%d")

                submit_btn = st.form_submit_button("Simpan Karyawan")
                if submit_btn:
                    clean_id = new_id.strip().upper()
                    existing_ids = (
                        [
                            str(x).strip().upper()
                            for x in st.session_state.employees["ID"].values
                        ]
                        if "ID" in st.session_state.employees.columns
                        else []
                    )

                    if not clean_id or not new_name or not new_role or not new_cc:
                        st.error("Mohon isi semua kolom yang wajib!")
                    elif clean_id in existing_ids:
                        st.error(f"❌ ID '{clean_id}' sudah digunakan!")
                    else:
                        new_row = {
                            "ID": clean_id,
                            "Nama Lengkap": new_name.strip().title(),
                            "Posisi": new_role.strip(),
                            "Cost Center": new_cc.strip(),
                            "Tanggal Bergabung": new_join.strftime("%Y-%m-%d"),
                            "Akhir Kontrak": new_end.strftime("%Y-%m-%d"),
                            "Tanggal Resign": new_resign_date,
                            "Site": new_site.strip(),
                            "Status": new_status,
                            "Terakhir Diperbarui": str(date.today()),
                        }
                        updated_df = pd.concat(
                            [
                                st.session_state.employees,
                                pd.DataFrame([new_row]),
                            ],
                            ignore_index=True,
                        )
                        save_data(updated_df)
                        st.success(
                            f"✅ ID '{clean_id}' berhasil ditambahkan!"
                        )
                        st.rerun()

        # 2. Bulk Import Data Master
        with st.sidebar.expander("📥 Import Banyak Data", expanded=False):
            import_type = st.radio(
                "Metode Import:", ["File CSV", "Tempel Teks (Excel/TSV)"]
            )

            if import_type == "File CSV":
                uploaded_file = st.file_uploader(
                    "Pilih file CSV", type=["csv"]
                )
                if uploaded_file is not None and st.button("Mulai Import File"):
                    try:
                        df_import = pd.read_csv(
                            uploaded_file, dtype={"ID": str}
                        )
                        df_import.columns = [
                            c.strip() for c in df_import.columns
                        ]
                        if "Jabatan" in df_import.columns:
                            df_import.rename(
                                columns={"Jabatan": "Posisi"}, inplace=True
                            )
                        if "Status" not in df_import.columns:
                            df_import["Status"] = "Aktif"
                        if "Tanggal Resign" not in df_import.columns:
                            df_import["Tanggal Resign"] = "-"
                        df_import["Terakhir Diperbarui"] = str(date.today())

                        existing_ids = set(
                            str(x).strip().upper()
                            for x in st.session_state.employees["ID"].values
                        )
                        df_import_filtered = df_import[
                            ~df_import["ID"]
                            .astype(str)
                            .str.strip()
                            .str.upper()
                            .isin(existing_ids)
                        ]
                        added_count = len(df_import_filtered)

                        if added_count > 0:
                            updated_df = pd.concat(
                                [
                                    st.session_state.employees,
                                    df_import_filtered,
                                ],
                                ignore_index=True,
                            )
                            save_data(updated_df)
                            st.success(
                                f"Berhasil mengimpor {added_count} data!"
                            )
                            st.rerun()
                        else:
                            st.error("Semua ID pada file sudah terdaftar!")
                    except Exception as e:
                        st.error(f"Gagal membaca file: {e}")
            else:
                pasted_text = st.text_area("Tempel dari Excel", height=150)
                if st.button("Mulai Import Teks") and pasted_text.strip():
                    lines = pasted_text.strip().split("\n")
                    added_rows = []
                    existing_ids = set(
                        str(x).strip().upper()
                        for x in st.session_state.employees["ID"].values
                    )

                    for line in lines:
                        delimiter = (
                            "\t"
                            if "\t" in line
                            else (";" if ";" in line else ",")
                        )
                        cols = [c.strip() for c in line.split(delimiter)]
                        if len(cols) >= 4:
                            emp_id, name, role_title, cc = (
                                cols[0].upper(),
                                cols[1].title(),
                                cols[2],
                                cols[3],
                            )
                            join_d = cols[4] if len(cols) > 4 else ""
                            end_d = cols[5] if len(cols) > 5 else ""
                            resign_d = cols[6] if len(cols) > 6 else "-"
                            site_val = cols[7] if len(cols) > 7 else ""
                            status_val = (
                                cols[8] if len(cols) > 8 else "Aktif"
                            )

                            if emp_id not in existing_ids:
                                added_rows.append({
                                    "ID": emp_id,
                                    "Nama Lengkap": name,
                                    "Posisi": role_title,
                                    "Cost Center": cc,
                                    "Tanggal Bergabung": join_d,
                                    "Akhir Kontrak": end_d,
                                    "Tanggal Resign": resign_d,
                                    "Site": site_val,
                                    "Status": status_val,
                                    "Terakhir Diperbarui": str(date.today()),
                                })
                                existing_ids.add(emp_id)

                    if added_rows:
                        updated_df = pd.concat(
                            [
                                st.session_state.employees,
                                pd.DataFrame(added_rows),
                            ],
                            ignore_index=True,
                        )
                        save_data(updated_df)
                        st.success(
                            f"Berhasil menambahkan {len(added_rows)} data baru!"
                        )
                        st.rerun()

        # 3. KUNCI & HAPUS DATA SNAPSHOT BULANAN
        with st.sidebar.expander(
            "📸 Freeze / Snapshot Bulanan", expanded=False
        ):
            st.subheader("🔒 Simpan Snapshot Baru")
            selected_periode = st.date_input(
                "Pilih Bulan Periode", value=date.today()
            ).strftime("%Y-%m")

            if st.button(f"🔒 Kunci Data {selected_periode}"):
                try:
                    df_curr = st.session_state.employees.copy()
                    df_active = (
                        df_curr[df_curr["Status"] == "Aktif"].copy()
                        if "Status" in df_curr.columns
                        else df_curr.copy()
                    )
                    df_active["Periode"] = selected_periode
                    df_active["Tanggal Snapshot"] = str(date.today())

                    cols_order = [
                        "Periode",
                        "ID",
                        "Nama Lengkap",
                        "Posisi",
                        "Cost Center",
                        "Tanggal Bergabung",
                        "Akhir Kontrak",
                        "Tanggal Resign",
                        "Site",
                        "Status",
                        "Terakhir Diperbarui",
                        "Tanggal Snapshot",
                    ]

                    df_old_snap = load_snapshot_data()
                    if (
                        not df_old_snap.empty
                        and "Periode" in df_old_snap.columns
                    ):
                        df_old_snap = df_old_snap[
                            df_old_snap["Periode"] != selected_periode
                        ]
                        df_new_snap = pd.concat([
                            df_old_snap,
                            df_active[cols_order],
                        ])
                    else:
                        df_new_snap = df_active[cols_order]

                    df_new_snap = df_new_snap.fillna("")
                    conn.update(
                        worksheet="Snapshot_Bulanan", data=df_new_snap
                    )
                    st.success(
                        f"✅ Rekap {selected_periode} berhasil disimpan!"
                    )
                    st.rerun()
                except Exception as e:
                    st.error(f"Gagal melakukan snapshot: {e}")

            st.markdown("---")
            st.subheader("🗑️ Hapus Snapshot Periode")
            df_snap_exist = load_snapshot_data()
            if (
                not df_snap_exist.empty
                and "Periode" in df_snap_exist.columns
            ):
                list_snap_periods = sorted(
                    df_snap_exist["Periode"].unique(), reverse=True
                )
                period_to_delete = st.selectbox(
                    "Pilih Periode yang Ingin Dihapus:", list_snap_periods
                )

                if st.button(f"🗑️ Hapus Snapshot {period_to_delete}"):
                    try:
                        df_snap_filtered = df_snap_exist[
                            df_snap_exist["Periode"] != period_to_delete
                        ]
                        df_snap_filtered = df_snap_filtered.fillna("")
                        conn.update(
                            worksheet="Snapshot_Bulanan", data=df_snap_filtered
                        )
                        st.success(
                            f"✅ Snapshot periode {period_to_delete} berhasil"
                            " dihapus!"
                        )
                        st.rerun()
                    except Exception as e:
                        st.error(f"Gagal menghapus snapshot: {e}")

        # Ekspor Database
        st.sidebar.markdown("---")
        st.sidebar.subheader("📤 Ekspor Database")
        csv_data = (
            st.session_state.employees.to_csv(index=False).encode("utf-8-sig")
        )
        st.sidebar.download_button(
            label="📄 Ekspor CSV",
            data=csv_data,
            file_name="ekspor_database_karyawan.csv",
            mime="text/csv",
            use_container_width=True,
        )
        excel_data = generate_excel_formatted(st.session_state.employees)
        st.sidebar.download_button(
            label="📊 Ekspor Excel Formatted (.xlsx)",
            data=excel_data,
            file_name=f"Rekap_Karyawan_{date.today().strftime('%Y%m%d')}.xlsx",
            mime=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            use_container_width=True,
        )

    # --- HALAMAN UTAMA MASTER ---
    if is_admin:
        st.info("🔓 **Mode Akses:** Administrator")
    else:
        st.info("👁️ **Mode Akses:** Umum / Guest (View Only)")

    # --- DASHBOARD ANALYTICS ---
    with st.expander(
        "📊 **Dashboard Analytics & Visualisasi Data**", expanded=True
    ):
        df_snap_hist = load_snapshot_data()

        current_period = date.today().strftime("%Y-%m")

        available_periods = []
        if not df_snap_hist.empty and "Periode" in df_snap_hist.columns:
            available_periods = sorted(
                list(df_snap_hist["Periode"].unique()), reverse=True
            )

        realtime_option = f"{current_period} (Bulan Berjalan - Realtime)"
        if realtime_option not in available_periods:
            available_periods.insert(0, realtime_option)

        selected_dash_period = st.selectbox(
            "📅 Pilih Periode Dashboard:", options=available_periods, index=0
        )

        if "Realtime" in selected_dash_period:
            df_ana = st.session_state.employees.copy()
            active_period_str = current_period
        else:
            df_ana = df_snap_hist[
                df_snap_hist["Periode"] == selected_dash_period
            ].copy()
            active_period_str = selected_dash_period

        def filter_resign_for_period(df, target_period):
            if df.empty or "Status" not in df.columns:
                return df

            mask_aktif = df["Status"].astype(str).str.strip().isin(["Aktif", "PKWT"])

            mask_resign_bulan_ini = pd.Series(False, index=df.index)
            if "Tanggal Resign" in df.columns:
                resign_dates = pd.to_datetime(
                    df["Tanggal Resign"], errors="coerce"
                ).dt.strftime("%Y-%m")
                mask_resign_bulan_ini = (
                    df["Status"].astype(str).str.strip() == "Resign"
                ) & (resign_dates == target_period)

            return df[mask_aktif | mask_resign_bulan_ini]

        df_pie_chart = filter_resign_for_period(df_ana, active_period_str)

        if not df_ana.empty:
            tab_overview, tab_trend, tab_cost = st.tabs([
                "📈 Ringkasan & Status",
                "🗓️ Tren Snapshot Bulanan",
                "💳 Sebaran Cost Center & Site",
            ])

            with tab_overview:
                c1, c2 = st.columns(2)
                with c1:
                    if (
                        "Status" in df_pie_chart.columns
                        and not df_pie_chart.empty
                    ):
                        fig_status = px.pie(
                            df_pie_chart,
                            names="Status",
                            title=(
                                "Komposisi Status Karyawan"
                                f" ({selected_dash_period})"
                            ),
                            hole=0.4,
                            color_discrete_sequence=px.colors.qualitative.Set2,
                        )
                        fig_status.update_traces(
                            textposition="inside", textinfo="percent+label"
                        )
                        st.plotly_chart(fig_status, use_container_width=True)
                    else:
                        st.info("Tidak ada data status untuk ditampilkan.")
                with c2:
                    if "Posisi" in df_ana.columns:
                        df_posisi_aktif = (
                            df_ana[df_ana["Status"] == "Aktif"]
                            if "Status" in df_ana.columns
                            else df_ana
                        )
                        top_roles = (
                            df_posisi_aktif["Posisi"]
                            .value_counts()
                            .head(10)
                            .reset_index()
                        )
                        top_roles.columns = ["Posisi", "Jumlah"]
                        fig_role = px.bar(
                            top_roles,
                            x="Jumlah",
                            y="Posisi",
                            orientation="h",
                            title=(
                                "Top 10 Posisi Terbanyak Karyawan Aktif"
                                f" ({selected_dash_period})"
                            ),
                            color="Jumlah",
                            color_continuous_scale="Blues",
                        )
                        fig_role.update_traces(textangle=0)
                        fig_role.update_layout(
                            yaxis={"categoryorder": "total ascending"}
                        )
                        st.plotly_chart(fig_role, use_container_width=True)

            with tab_trend:
                if (
                    not df_snap_hist.empty
                    and "Periode" in df_snap_hist.columns
                ):
                    trend_summary = (
                        df_snap_hist.groupby("Periode")["ID"]
                        .count()
                        .reset_index(name="Karyawan Aktif")
                        .sort_values("Periode")
                    )
                    fig_trend = px.line(
                        trend_summary,
                        x="Periode",
                        y="Karyawan Aktif",
                        markers=True,
                        title=(
                            "Pertumbuhan Jumlah Karyawan Aktif per Periode"
                            " Snapshot"
                        ),
                        line_shape="spline",
                    )
                    fig_trend.update_traces(
                        line_color="#1F4E79", line_width=3, marker_size=8
                    )
                    st.plotly_chart(fig_trend, use_container_width=True)
                else:
                    st.info("Belum ada data snapshot historis.")

            with tab_cost:
                c3, c4 = st.columns(2)
                with c3:
                    if "Cost Center" in df_ana.columns:
                        df_cc_clean = df_ana.copy()
                        df_cc_clean["Cost Center Clean"] = (
                            df_cc_clean["Cost Center"]
                            .astype(str)
                            .str.strip()
                            .str.title()
                            .replace("", "Belum Diisi")
                        )
                        df_cc_clean["Cost Center Clean"] = df_cc_clean[
                            "Cost Center Clean"
                        ].replace({
                            "Vinfast": "VinFast",
                            "Cj Food": "CJ Food",
                            "Fks": "FKS",
                            "Keva & Jotun": "Keva & Jotun",
                            "Jotun, Keva": "Keva & Jotun",
                        })
                        cc_counts = (
                            df_cc_clean["Cost Center Clean"]
                            .value_counts()
                            .reset_index()
                        )
                        cc_counts.columns = ["Cost Center", "Jumlah"]
                        fig_cc = px.bar(
                            cc_counts,
                            x="Jumlah",
                            y="Cost Center",
                            orientation="h",
                            title=(
                                "Jumlah Karyawan per Cost Center"
                                f" ({selected_dash_period})"
                            ),
                            color="Jumlah",
                            color_continuous_scale="Viridis",
                            text="Jumlah",
                        )
                        fig_cc.update_traces(
                            textposition="outside", textangle=0
                        )
                        fig_cc.update_layout(
                            yaxis={"categoryorder": "total ascending"},
                            height=max(450, len(cc_counts) * 25),
                        )
                        st.plotly_chart(fig_cc, use_container_width=True)
                with c4:
                    if "Site" in df_ana.columns:
                        df_site_clean = df_ana.copy()
                        df_site_clean["Site Clean"] = (
                            df_site_clean["Site"]
                            .astype(str)
                            .str.strip()
                            .str.upper()
                            .replace("", "BELUM DIISI")
                        )
                        site_counts = (
                            df_site_clean["Site Clean"]
                            .value_counts()
                            .reset_index()
                        )
                        site_counts.columns = ["Site", "Jumlah"]
                        fig_site = px.pie(
                            site_counts,
                            names="Site",
                            values="Jumlah",
                            title=(
                                "Distribusi Lokasi Kerja / Site"
                                f" ({selected_dash_period})"
                            ),
                            hole=0.3,
                        )
                        st.plotly_chart(fig_site, use_container_width=True)

    st.divider()

    # --- FITUR PENCARIAN & FILTER MASTER ---
    col_mode, col_cat, col_src = st.columns([1.5, 1.5, 3])
    with col_mode:
        view_mode = st.selectbox(
            "Tampilkan Data:",
            ["Master Real-time", "Rekap Snapshot Bulanan"],
        )

    df_display = pd.DataFrame()
    if view_mode == "Rekap Snapshot Bulanan":
        df_snap_all = load_snapshot_data()
        if not df_snap_all.empty and "Periode" in df_snap_all.columns:
            list_periode = sorted(
                df_snap_all["Periode"].unique(), reverse=True
            )
            selected_view_period = st.selectbox(
                "Pilih Periode Rekap:", list_periode
            )
            df_display = df_snap_all[
                df_snap_all["Periode"] == selected_view_period
            ].copy()
        else:
            st.warning("Belum ada data snapshot yang disimpan.")
    else:
        df_display = st.session_state.employees.copy()

    with col_cat:
        search_category = st.selectbox(
            "Cari Berdasarkan:",
            [
                "Semua Kolom",
                "Nama Lengkap",
                "Posisi",
                "Cost Center",
                "Site",
                "Status",
            ],
        )

    with col_src:
        search_query = st.text_input("🔍 Masukkan kata kunci pencarian...", "")

    if search_query and not df_display.empty:
        query = search_query.strip().lower()
        if search_category in df_display.columns:
            df_display = df_display[
                df_display[search_category]
                .astype(str)
                .str.lower()
                .str.contains(query, na=False)
            ]
        else:
            mask = pd.Series(False, index=df_display.index)
            for col in ["Nama Lengkap", "Posisi", "Cost Center", "Site", "Status"]:
                if col in df_display.columns:
                    mask |= (
                        df_display[col]
                        .astype(str)
                        .str.lower()
                        .str.contains(query, na=False)
                    )
            df_display = df_display[mask]

    col_tb_title, col_pdf_btn = st.columns([3, 1])
    with col_tb_title:
        st.subheader(f"📋 Tabel Data Karyawan ({view_mode})")
    with col_pdf_btn:
        if not df_display.empty:
            pdf_bytes = generate_pdf(df_display)
            st.download_button(
                label="📄 Cetak / Download PDF",
                data=pdf_bytes,
                file_name=(
                    f"Laporan_Karyawan_{date.today().strftime('%Y%m%d')}.pdf"
                ),
                mime="application/pdf",
                use_container_width=True,
            )

    if df_display.empty:
        st.warning("Tidak ada data karyawan yang cocok dengan pencarian.")
    else:
        st.dataframe(df_display, use_container_width=True)

    # Edit Data
    if (
        is_admin
        and view_mode == "Master Real-time"
        and not st.session_state.employees.empty
    ):
        st.divider()
        st.subheader("🛠️ Kelola / Edit / Ubah Status Data Karyawan")
        selected_id = st.selectbox(
            "Pilih ID Karyawan untuk Diubah / Dihapus:",
            options=["-- Pilih ID --"] + list(st.session_state.employees["ID"]),
        )

        if selected_id != "-- Pilih ID --":
            emp_idx = st.session_state.employees[
                st.session_state.employees["ID"] == selected_id
            ].index[0]
            row = st.session_state.employees.loc[emp_idx]

            with st.form("edit_form"):
                st.write(
                    f"Editing: **{row['Nama Lengkap']}** (ID: `{row['ID']}`)"
                )
                e_name = st.text_input("Nama Lengkap", value=row["Nama Lengkap"])
                e_role = st.text_input("Posisi", value=row.get("Posisi", ""))
                e_cc = st.text_input("Cost Center", value=row["Cost Center"])
                e_join = st.text_input(
                    "Tanggal Bergabung (YYYY-MM-DD)",
                    value=row["Tanggal Bergabung"],
                )
                e_end = st.text_input(
                    "Akhir Kontrak (YYYY-MM-DD)", value=row["Akhir Kontrak"]
                )
                e_site = st.text_input(
                    "Site / Lokasi Kerja", value=row.get("Site", "")
                )
                current_status = row.get("Status", "Aktif")
                status_opts = ["Aktif", "Resign", "PKWT"]
                idx_stat = (
                    status_opts.index(current_status)
                    if current_status in status_opts
                    else 0
                )
                e_status = st.selectbox(
                    "Status Karyawan", options=status_opts, index=idx_stat
                )
                e_resign = st.text_input(
                    "Tanggal Resign (YYYY-MM-DD)",
                    value=row.get("Tanggal Resign", "-"),
                )

                col_save, col_del = st.columns(2)
                with col_save:
                    btn_save = st.form_submit_button("💾 Simpan Perubahan")
                with col_del:
                    btn_del = st.form_submit_button("🗑️ Hapus Karyawan")

                if btn_save:
                    st.session_state.employees.loc[
                        emp_idx,
                        [
                            "Nama Lengkap",
                            "Posisi",
                            "Cost Center",
                            "Tanggal Bergabung",
                            "Akhir Kontrak",
                            "Tanggal Resign",
                            "Site",
                            "Status",
                            "Terakhir Diperbarui",
                        ],
                    ] = [
                        e_name.strip().title(),
                        e_role.strip(),
                        e_cc.strip(),
                        e_join.strip(),
                        e_end.strip(),
                        e_resign.strip(),
                        e_site.strip(),
                        e_status,
                        str(date.today()),
                    ]
                    save_data(st.session_state.employees)
                    st.success("Data berhasil diperbarui!")
                    st.rerun()

                if btn_del:
                    updated_df = st.session_state.employees.drop(
                        emp_idx
                    ).reset_index(drop=True)
                    save_data(updated_df)
                    st.success("Data karyawan berhasil dihapus!")
                    st.rerun()


# ==============================================================================
# MODUL 2: REKAP ABSENSI (TIMESHEET MATRIX)
# ==============================================================================
if menu_pilihan == "⏱️ Rekap Absensi (Timesheet)":

    st.title("⏱️ Rekap & Import Absensi Karyawan Site")
    st.caption(
        "Upload file Excel Timesheet untuk memperbarui rekap absensi di Google"
        " Sheets."
    )

    def load_absensi_data():
        try:
            df_absen = conn.read(worksheet="Absensi_Karyawan", ttl=0)
            if df_absen is not None and not df_absen.empty:
                df_absen["ID"] = (
                    df_absen["ID"].astype(str).str.strip().str.upper()
                )
                df_absen["Nama Lengkap"] = (
                    df_absen["Nama Lengkap"]
                    .astype(str)
                    .str.strip()
                    .str.title()
                )
                df_absen["Tanggal"] = pd.to_datetime(
                    df_absen["Tanggal"]
                ).dt.strftime("%Y-%m-%d")
                if "Status" not in df_absen.columns:
                    df_absen["Status"] = "Hadir"
            return df_absen if df_absen is not None else pd.DataFrame()
        except Exception:
            return pd.DataFrame(
                columns=[
                    "ID",
                    "Nama Lengkap",
                    "Site",
                    "Job Title",
                    "Tanggal",
                    "In",
                    "Out",
                    "Shift",
                    "Status",
                ]
            )

    if "df_absensi" not in st.session_state or st.sidebar.button(
        "🔄 Refresh Data Absensi"
    ):
        st.session_state.df_absensi = load_absensi_data()

    if is_admin:
        with st.expander(
            "📥 **Upload File Excel Timesheet**", expanded=False
        ):
            st.info(
                "Pastikan file Excel memiliki 9 kolom: ID, Nama Lengkap, Site,"
                " Job Title, Tanggal, In, Out, Shift (atau Sta), Status"
            )
            uploaded_file = st.file_uploader(
                "Pilih File Excel:", type=["xlsx", "xls"]
            )

            if uploaded_file is not None and st.button(
                "🚀 Simpan ke Database Google Sheets"
            ):
                try:
                    df_upload = pd.read_excel(uploaded_file)
                    df_upload.columns = [c.strip() for c in df_upload.columns]

                    df_upload.rename(
                        columns={"Sta": "Shift", "Ket": "Status"}, inplace=True
                    )
                    if "Status" not in df_upload.columns:
                        df_upload["Status"] = "Hadir"

                    df_upload["Tanggal"] = pd.to_datetime(
                        df_upload["Tanggal"]
                    ).dt.strftime("%Y-%m-%d")
                    df_upload["ID"] = (
                        df_upload["ID"].astype(str).str.strip().str.upper()
                    )
                    df_upload["Nama Lengkap"] = (
                        df_upload["Nama Lengkap"]
                        .astype(str)
                        .str.strip()
                        .str.title()
                    )

                    df_lama = load_absensi_data()
                    updated_absensi = pd.concat(
                        [df_lama, df_upload], ignore_index=True
                    )
                    updated_absensi = updated_absensi.drop_duplicates(
                        subset=["ID", "Tanggal"], keep="last"
                    )

                    clean_absensi = updated_absensi.fillna("")
                    conn.update(
                        worksheet="Absensi_Karyawan", data=clean_absensi
                    )
                    st.session_state.df_absensi = clean_absensi

                    st.success(
                        f"✅ Berhasil menyimpan {len(df_upload)} baris data"
                        " absensi ke Google Sheets!"
                    )
                    st.rerun()
                except Exception as e:
                    st.error(f"Gagal memproses file Excel: {e}")

    df_absen = st.session_state.df_absensi

    if not df_absen.empty:
        with st.expander(
            "📊 **Dashboard Analytics & Visualisasi Data Absensi**",
            expanded=True,
        ):
            df_analytics = df_absen.copy()

            def clean_status_val(row):
                status_raw = str(row.get("Status", "")).strip().lower()
                in_val = str(row.get("In", "")).strip().lower()
                out_val = str(row.get("Out", "")).strip().lower()

                if status_raw in ["sakit"]:
                    return "Sakit"
                elif status_raw in ["cuti"]:
                    return "Cuti"
                elif status_raw in ["izin", "ijin"]:
                    return "Izin"
                elif status_raw in ["alpha", "mangkir", "tidak hadir"]:
                    return "Tidak Hadir"
                elif status_raw in ["late", "terlambat"]:
                    return "Late"

                in_empty = pd.isna(row.get("In")) or in_val in [
                    "none",
                    "nan",
                    "",
                    "-",
                    "null",
                ]
                out_empty = pd.isna(row.get("Out")) or out_val in [
                    "none",
                    "nan",
                    "",
                    "-",
                    "null",
                ]

                if in_empty and out_empty:
                    return "Tidak Hadir"

                return "Hadir"

            df_analytics["Status_Clean"] = df_analytics.apply(
                clean_status_val, axis=1
            )
            df_analytics["Tanggal_Clean"] = pd.to_datetime(
                df_analytics["Tanggal"]
            ).dt.strftime("%Y-%m-%d")

            total_records = len(df_analytics)
            hadir_count = len(
                df_analytics[df_analytics["Status_Clean"] == "Hadir"]
            )
            late_count = len(
                df_analytics[df_analytics["Status_Clean"] == "Late"]
            )
            tidak_hadir_count = len(
                df_analytics[
                    df_analytics["Status_Clean"].isin([
                        "Sakit",
                        "Cuti",
                        "Izin",
                        "Tidak Hadir",
                    ])
                ]
            )

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total Record Absensi", f"{total_records:,}")
            m2.metric(
                "Total Hadir Normal",
                f"{hadir_count:,}",
                delta=(
                    f"{round(hadir_count/total_records*100, 1) if total_records else 0}%"
                ),
            )
            m3.metric(
                "Terlambat (Late)",
                f"{late_count:,}",
                delta=f"-{late_count}" if late_count > 0 else "0",
                delta_color="inverse",
            )
            m4.metric(
                "Tidak Hadir (Sakit/Cuti/Izin/Tidak Hadir)",
                f"{tidak_hadir_count:,}",
            )

            st.markdown("---")

            tab_stat, tab_shift, tab_top_late = st.tabs([
                "📊 Ringkasan Status",
                "⏱️ Sebaran Shift Work",
                "⚠️ Catatan Status Khusus",
            ])

            with tab_stat:
                c1, c2 = st.columns(2)
                with c1:
                    status_counts = (
                        df_analytics["Status_Clean"]
                        .value_counts()
                        .reset_index()
                    )
                    status_counts.columns = ["Status", "Jumlah"]
                    fig_status = px.pie(
                        status_counts,
                        names="Status",
                        values="Jumlah",
                        title="Komposisi Status Kehadiran Karyawan",
                        hole=0.4,
                        color_discrete_map={
                            "Hadir": "#66C2A5",
                            "Sakit": "#FFC000",
                            "Cuti": "#1F4E79",
                            "Izin": "#17BECF",
                            "Late": "#FC8D62",
                            "Tidak Hadir": "#E78AC3",
                        },
                    )
                    fig_status.update_traces(
                        textposition="inside", textinfo="percent+label"
                    )
                    st.plotly_chart(fig_status, use_container_width=True)

                with c2:
                    daily_trend = (
                        df_analytics.groupby(["Tanggal_Clean", "Status_Clean"])[
                            "ID"
                        ]
                        .count()
                        .reset_index(name="Total Scan")
                    )
                    fig_daily = px.bar(
                        daily_trend,
                        x="Tanggal_Clean",
                        y="Total Scan",
                        color="Status_Clean",
                        title="Volume Absensi Harian (Berdasarkan Status)",
                        text="Total Scan",
                        color_discrete_map={
                            "Hadir": "#1F4E79",
                            "Sakit": "#FFC000",
                            "Cuti": "#2CA02C",
                            "Izin": "#17BECF",
                            "Late": "#FF7F0E",
                            "Tidak Hadir": "#D62728",
                        },
                    )
                    fig_daily.update_traces(textangle=0)
                    fig_daily.update_xaxes(
                        type="category", title_text="Tanggal"
                    )
                    fig_daily.update_layout(legend_title_text="Status")
                    st.plotly_chart(fig_daily, use_container_width=True)

            with tab_shift:
                if (
                    "Shift" in df_analytics.columns
                    and "Tanggal_Clean" in df_analytics.columns
                ):
                    shift_df = df_analytics.copy()
                    shift_df["Shift_Clean"] = (
                        shift_df["Shift"]
                        .astype(str)
                        .str.replace(".0", "", regex=False)
                        .str.strip()
                        .str.upper()
                    )

                    target_shifts = ["1", "2", "3", "M"]
                    shift_filtered = shift_df[
                        shift_df["Shift_Clean"].isin(target_shifts)
                    ]

                    if not shift_filtered.empty:
                        total_days = shift_df["Tanggal_Clean"].nunique()
                        shift_avg = (
                            shift_filtered.groupby("Shift_Clean")["ID"]
                            .count()
                            .div(total_days)
                            .round(1)
                            .reset_index(name="Rata_Rata_Karyawan")
                        )

                        shift_order = {"1": 1, "2": 2, "3": 3, "M": 4}
                        shift_avg["Order"] = shift_avg["Shift_Clean"].map(
                            shift_order
                        )
                        shift_avg = shift_avg.sort_values("Order")

                        fig_shift = px.bar(
                            shift_avg,
                            x="Shift_Clean",
                            y="Rata_Rata_Karyawan",
                            text="Rata_Rata_Karyawan",
                            title=(
                                "Rata-Rata Jumlah Karyawan per Hari (Total"
                                f" {total_days} Hari Data)"
                            ),
                            color="Rata_Rata_Karyawan",
                            color_continuous_scale="Viridis",
                            labels={
                                "Shift_Clean": "Shift Work",
                                "Rata_Rata_Karyawan": "Rata-Rata Orang / Hari",
                            },
                        )
                        fig_shift.update_traces(
                            textposition="outside",
                            texttemplate="%{text} orang/hari",
                            textangle=0,
                        )
                        fig_shift.update_xaxes(type="category")
                        st.plotly_chart(fig_shift, use_container_width=True)
                    else:
                        st.info(
                            "Tidak ditemukan data untuk Shift 1, 2, 3, atau"
                            " Middle (M)."
                        )

            with tab_top_late:
                df_late_only = df_analytics[
                    df_analytics["Status_Clean"].isin([
                        "Late",
                        "Terlambat",
                        "Sakit",
                        "Cuti",
                        "Izin",
                        "Tidak Hadir",
                    ])
                ]
                if not df_late_only.empty:
                    top_late = (
                        df_late_only.groupby(["Nama Lengkap", "Status_Clean"])
                        .size()
                        .reset_index(name="Frekuensi")
                    )
                    top_employees = (
                        top_late.groupby("Nama Lengkap")["Frekuensi"]
                        .sum()
                        .nlargest(10)
                        .index
                    )
                    top_late = top_late[
                        top_late["Nama Lengkap"].isin(top_employees)
                    ]

                    fig_top = px.bar(
                        top_late,
                        x="Frekuensi",
                        y="Nama Lengkap",
                        color="Status_Clean",
                        orientation="h",
                        title=(
                            "Top 10 Karyawan Catatan Khusus (Rincian per"
                            " Status)"
                        ),
                        text="Frekuensi",
                        color_discrete_map={
                            "Late": "#FF0000",
                            "Sakit": "#FFC000",
                            "Cuti": "#1F4E79",
                            "Izin": "#17BECF",
                            "Tidak Hadir": "#8B0000",
                        },
                    )
                    fig_top.update_traces(textangle=0)
                    fig_top.update_layout(
                        yaxis={"categoryorder": "total ascending"},
                        legend_title_text="Status Khusus",
                    )
                    st.plotly_chart(fig_top, use_container_width=True)
                else:
                    st.success(
                        "🎉 Tidak ditemukan catatan keterlambatan atau"
                        " ketidakhadiran khusus pada data absensi saat ini."
                    )

    st.divider()
    st.subheader("📊 Timesheet Matrix")

    if df_absen.empty:
        st.warning(
            "Belum ada data absensi di Google Sheets. Silakan upload file Excel"
            " terlebih dahulu."
        )
    else:
        list_site = ["Semua Site"] + sorted(
            list(df_absen["Site"].dropna().astype(str).unique())
        )
        selected_site = st.selectbox("Tampilkan Site:", list_site)

        if selected_site != "Semua Site":
            df_absen = df_absen[df_absen["Site"] == selected_site]

        df_absen_clean = df_absen.copy()
        df_absen_clean["ID"] = (
            df_absen_clean["ID"].astype(str).str.strip().str.upper()
        )
        df_absen_clean["Nama Lengkap"] = (
            df_absen_clean["Nama Lengkap"]
            .astype(str)
            .str.strip()
            .str.title()
        )

        id_to_name = (
            df_absen_clean.groupby("ID")["Nama Lengkap"].last().to_dict()
        )
        df_absen_clean["Nama Lengkap"] = df_absen_clean["ID"].map(id_to_name)

        df_absen_clean = df_absen_clean.sort_values(
            by=["ID", "Tanggal", "In"], ascending=[True, True, False]
        )
        df_absen_clean = df_absen_clean.drop_duplicates(
            subset=["ID", "Tanggal"], keep="first"
        ).copy()

        df_absen_clean["Tgl_Format"] = pd.to_datetime(
            df_absen_clean["Tanggal"]
        ).dt.strftime("%d-%b\n%a")

        def clean_shift(val):
            if pd.isna(val) or str(val).strip().lower() in [
                "none",
                "nan",
                "",
                "-",
            ]:
                return "-"
            try:
                val_float = float(val)
                if val_float.is_integer():
                    return str(int(val_float))
                return str(val)
            except ValueError:
                return str(val)

        df_absen_clean["Shift"] = df_absen_clean["Shift"].apply(clean_shift)

        df_melted = df_absen_clean.melt(
            id_vars=["ID", "Nama Lengkap", "Tgl_Format"],
            value_vars=["In", "Out", "Shift", "Status"],
            var_name="SubHeader",
            value_name="Value",
        )

        matrix_df = df_melted.pivot_table(
            index=["ID", "Nama Lengkap"],
            columns=["Tgl_Format", "SubHeader"],
            values="Value",
            aggfunc="first",
        )

        # --- PERBAIKAN TIMESHET MATRIX (TETAP MENJAGA KOLOM STATUS KOSONG) ---
        unique_dates = df_absen_clean["Tgl_Format"].unique()
        sub_headers = ["In", "Out", "Shift", "Status"]

        # Buat struktur MultiIndex secara eksplisit untuk semua tanggal
        full_columns = pd.MultiIndex.from_product(
            [unique_dates, sub_headers], names=["Tgl_Format", "SubHeader"]
        )

        # Paksa reindex seluruh struktur kolom agar 'Status' tidak hilang
        matrix_df = matrix_df.reindex(columns=full_columns)

        matrix_df = matrix_df.fillna("-")
        matrix_df = matrix_df.map(
            lambda x: (
                "-" if str(x).strip().lower() in ["none", "nan", ""] else x
            )
        )

        def apply_matrix_styles(df):
            styles_df = pd.DataFrame("", index=df.index, columns=df.columns)

            for col in df.columns:
                sub_header = col[1]

                if sub_header == "Status":
                    for idx in df.index:
                        val_str = str(df.loc[idx, col]).strip().lower()
                        if val_str in ["sakit", "cuti", "izin", "ijin"]:
                            styles_df.loc[idx, col] = (
                                "background-color: #FFC000; color: black;"
                                " font-weight: bold;"
                            )
                        elif val_str in ["late", "terlambat"]:
                            styles_df.loc[idx, col] = (
                                "background-color: #FF0000; color: white;"
                                " font-weight: bold;"
                            )
                        elif val_str in [
                            "alpha",
                            "mangkir",
                            "tidak hadir",
                        ]:
                            styles_df.loc[idx, col] = (
                                "background-color: #8B0000; color: white;"
                                " font-weight: bold;"
                            )

            return styles_df

        styled_matrix = matrix_df.style.apply(
            apply_matrix_styles, axis=None
        ).set_properties(
            **{
                "text-align": "center",
                "font-size": "12px",
                "border": "1px solid #d3d3d3",
            }
        )

        st.dataframe(styled_matrix, use_container_width=True, height=500)


# ==============================================================================
# MODUL 3: MANPOWER COST MANAGER (DIPROTEKSI PASSWORD)
# ==============================================================================
if menu_pilihan == "💳 Manpower Cost Manager":

    if not check_manpower_access():
        st.stop()

    col_mc_head, col_mc_lock = st.columns([4, 1])
    with col_mc_head:
        st.title("💳 Manpower Cost Manager")
        st.caption(
            "Modul Pengelolaan Biaya Tenaga Kerja (Manpower Cost), PPN/PPh 23, &"
            " Invoice Detail."
        )
    with col_mc_lock:
        st.write("")
        if st.button(
            "🔒 Kunci Modul",
            use_container_width=True,
            help="Keluar dan kunci kembali halaman ini",
        ):
            st.session_state["manpower_authenticated"] = False
            st.rerun()

    MANPOWER_COST_HEADERS = [
        "Month",
        "Invoice No",
        "Name",
        "Employee ID (by Vendor)",
        "Cost Center Name",
        "Department",
        "Work Location",
        "Job Position",
        "Type",
        "Gender",
        "Contract",
        "Employment Status",
        "Project",
        "Basic Salary",
        "Meals & Transp",
        "Overtime",
        "Position",
        "Skill",
        "Other",
        "Shortage",
        "Deduction",
        "Total Salary",
        "BPJS",
        "Management Fee",
        "Total Manpower Cost",
        "Grand Total",
        "CJI",
        "KHQ Report",
        "PPN",
        "PPH23",
        "Total Payment Amount",
    ]

    def load_manpower_cost_data():
        df_mc = None
        try:
            df_mc = conn.read(worksheet="Manpower_Cost", ttl=0)
        except Exception:
            try:
                df_mc = conn.read(ttl=0)
            except Exception:
                df_mc = None

        if (
            df_mc is not None
            and isinstance(df_mc, pd.DataFrame)
            and not df_mc.empty
        ):
            df_mc.columns = [str(c).strip() for c in df_mc.columns]
            col_map = {str(c).strip().lower(): c for c in df_mc.columns}

            new_df = pd.DataFrame()
            for target_col in MANPOWER_COST_HEADERS:
                key = target_col.strip().lower()
                if key in col_map:
                    new_df[target_col] = df_mc[col_map[key]]
                else:
                    new_df[target_col] = ""

            if "Month" in new_df.columns:
                new_df = new_df[new_df["Month"].astype(str).str.strip() != ""]

            return new_df

        return pd.DataFrame(columns=MANPOWER_COST_HEADERS)

    if "df_manpower_cost" not in st.session_state or st.sidebar.button(
        "🔄 Refresh Data Manpower Cost"
    ):
        st.session_state.df_manpower_cost = load_manpower_cost_data()

    df_mc = st.session_state.df_manpower_cost

    if is_admin:
        with st.expander(
            "📥 **Upload / Import File Manpower Cost**", expanded=False
        ):
            st.info(
                "Upload file Excel/CSV Manpower Cost bulanan untuk disimpan ke"
                " Google Sheets."
            )
            mc_file = st.file_uploader(
                "Pilih File Excel/CSV Manpower Cost:",
                type=["xlsx", "xls", "csv"],
                key="mc_uploader",
            )
            if mc_file is not None and st.button(
                "🚀 Simpan Data Manpower Cost ke Sheets"
            ):
                try:
                    if mc_file.name.endswith(".csv"):
                        df_mc_upload = pd.read_csv(mc_file)
                    else:
                        df_mc_upload = pd.read_excel(mc_file)

                    df_mc_upload.columns = [
                        str(c).strip() for c in df_mc_upload.columns
                    ]

                    for col in MANPOWER_COST_HEADERS:
                        if col not in df_mc_upload.columns:
                            df_mc_upload[col] = ""

                    df_mc_upload = df_mc_upload[MANPOWER_COST_HEADERS]

                    df_mc_old = load_manpower_cost_data()
                    updated_mc = pd.concat(
                        [df_mc_old, df_mc_upload], ignore_index=True
                    )
                    clean_mc = updated_mc.fillna("")

                    try:
                        conn.update(worksheet="Manpower_Cost", data=clean_mc)
                    except Exception:
                        conn.update(data=clean_mc)

                    st.session_state.df_manpower_cost = clean_mc
                    st.success(
                        f"✅ Berhasil mengimpor {len(df_mc_upload)} baris data"
                        " Manpower Cost!"
                    )
                    st.rerun()
                except Exception as e:
                    st.error(f"Gagal memproses file: {e}")

    st.divider()
    if df_mc.empty:
        st.info(
            "Sheet `Manpower_Cost` masih kosong. Silakan upload file terlebih"
            " dahulu."
        )
        st.dataframe(
            pd.DataFrame(columns=MANPOWER_COST_HEADERS),
            use_container_width=True,
        )
    else:
        df_mc_clean = df_mc.fillna("").astype(str)

        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            months = sorted([
                x.strip()
                for x in df_mc_clean["Month"].unique()
                if x.strip() != "" and x.strip().lower() != "nan"
            ])
            selected_months = st.multiselect(
                "Filter Bulan (Month):", options=months, default=months
            )
        with col_f2:
            projects = sorted([
                x.strip()
                for x in df_mc_clean["Project"].unique()
                if x.strip() != "" and x.strip().lower() != "nan"
            ])
            selected_project = st.multiselect(
                "Filter Project:", options=projects, default=projects
            )
        with col_f3:
            locations = sorted([
                x.strip()
                for x in df_mc_clean["Work Location"].unique()
                if x.strip() != "" and x.strip().lower() != "nan"
            ])
            selected_location = st.multiselect(
                "Filter Work Location:", options=locations, default=locations
            )

        filtered_mc = df_mc_clean.copy()
        if selected_months:
            filtered_mc = filtered_mc[
                filtered_mc["Month"].str.strip().isin(selected_months)
            ]
        if selected_project:
            filtered_mc = filtered_mc[
                filtered_mc["Project"].str.strip().isin(selected_project)
            ]
        if selected_location:
            filtered_mc = filtered_mc[
                filtered_mc["Work Location"].str.strip().isin(selected_location)
            ]

        def to_num(series):
            def parse_val(val):
                if pd.isna(val):
                    return 0.0
                s = str(val).strip()
                if not s or s.lower() in ["nan", "none", "-", ""]:
                    return 0.0

                s_clean = re.sub(r"[^\d.,-]", "", s)
                if not s_clean:
                    return 0.0

                if "," in s_clean and "." in s_clean:
                    if s_clean.rfind(",") > s_clean.rfind("."):
                        s_clean = s_clean.replace(".", "").replace(",", ".")
                    else:
                        s_clean = s_clean.replace(",", "")
                elif "," in s_clean:
                    parts = s_clean.split(",")
                    if len(parts) == 2 and len(parts[1]) <= 2:
                        s_clean = s_clean.replace(",", ".")
                    else:
                        s_clean = s_clean.replace(",", "")
                elif "." in s_clean:
                    parts = s_clean.split(".")
                    if len(parts) > 2:
                        s_clean = s_clean.replace(".", "")
                    elif len(parts) == 2 and len(parts[1]) == 3:
                        s_clean = s_clean.replace(".", "")

                try:
                    return float(s_clean)
                except ValueError:
                    return 0.0

            return series.apply(parse_val)

        total_headcount = len(filtered_mc)
        total_salary = int(
            round(to_num(filtered_mc["Total Salary"]).astype(float).sum())
        )
        total_mp_cost = int(
            round(
                to_num(filtered_mc["Total Manpower Cost"])
                .astype(float)
                .sum()
            )
        )
        total_payment = int(
            round(
                to_num(filtered_mc["Total Payment Amount"])
                .astype(float)
                .sum()
            )
        )

        km1, km2, km3, km4 = st.columns(4)
        km1.metric("Total Headcount", f"{total_headcount:,}")
        km2.metric("Total Salary", f"Rp {total_salary:,}".replace(",", "."))
        km3.metric(
            "Total Manpower Cost", f"Rp {total_mp_cost:,}".replace(",", ".")
        )
        km4.metric(
            "Total Payment Amount", f"Rp {total_payment:,}".replace(",", ".")
        )

        with st.expander(
            "📊 **Dashboard Analytics & Sebaran Biaya Manpower**", expanded=True
        ):
            df_chart = filtered_mc.copy()
            df_chart["Parsed_Payment"] = to_num(
                df_chart["Total Payment Amount"]
            )
            df_chart["Parsed_Salary"] = to_num(df_chart["Total Salary"])
            df_chart["Parsed_Overtime"] = to_num(df_chart["Overtime"])

            if not df_chart.empty:
                top_proj_name = (
                    df_chart.groupby("Project")["Parsed_Payment"]
                    .sum()
                    .idxmax()
                )
                top_proj_val = (
                    df_chart.groupby("Project")["Parsed_Payment"]
                    .sum()
                    .max()
                )
                formatted_top_val = (
                    f"Rp {int(top_proj_val):,}".replace(",", ".")
                )
                st.info(
                    f"💡 **Ringkasan Eksekutif:** Anggaran total payment terbesar"
                    f" saat ini dipegang oleh project **{top_proj_name}**"
                    f" dengan nilai sebesar **{formatted_top_val}**."
                )

            gc1, gc2 = st.columns(2)
            with gc1:
                trend_month = (
                    df_chart.groupby("Month")["Parsed_Payment"]
                    .sum()
                    .reset_index()
                )
                trend_month["Month_Dt"] = pd.to_datetime(
                    trend_month["Month"], errors="coerce"
                )
                trend_month = trend_month.sort_values(
                    by=["Month_Dt", "Month"], ascending=[True, True]
                )
                trend_month["Text_Format"] = trend_month[
                    "Parsed_Payment"
                ].apply(format_rp_short)

                fig_trend = px.bar(
                    trend_month,
                    x="Month",
                    y="Parsed_Payment",
                    title="Total Payment Amount per Bulan",
                    text="Text_Format",
                    color="Month",
                    color_discrete_sequence=["#2CA02C", "#1F4E79"],
                )
                fig_trend.update_traces(
                    textangle=0,
                    textposition="outside",
                    hovertemplate=(
                        "<b>Bulan:</b> %{x}<br><b>Total Payment:</b> Rp"
                        " %{y:,.0f}<extra></extra>"
                    ),
                )
                fig_trend.update_layout(
                    xaxis_title="Bulan",
                    yaxis_title="Total Payment (Rp)",
                    showlegend=False,
                    xaxis={
                        "categoryorder": "array",
                        "categoryarray": trend_month["Month"].tolist(),
                    },
                )
                st.plotly_chart(fig_trend, use_container_width=True)

            with gc2:
                top_cost_proj = (
                    df_chart.groupby("Project")["Parsed_Payment"]
                    .sum()
                    .nlargest(10)
                    .reset_index()
                )
                top_cost_proj["Text_Format"] = top_cost_proj[
                    "Parsed_Payment"
                ].apply(format_rp_short)

                fig_proj_cost = px.bar(
                    top_cost_proj,
                    x="Parsed_Payment",
                    y="Project",
                    orientation="h",
                    title="Top 10 Project Berdasarkan Total Payment",
                    text="Text_Format",
                    color="Parsed_Payment",
                    color_continuous_scale="Viridis",
                )
                fig_proj_cost.update_traces(
                    textangle=0,
                    textposition="outside",
                    hovertemplate=(
                        "<b>Project:</b> %{y}<br><b>Total Payment:</b> Rp"
                        " %{x:,.0f}<extra></extra>"
                    ),
                )
                fig_proj_cost.update_layout(
                    yaxis={"categoryorder": "total ascending"},
                    xaxis_title="Total Payment (Rp)",
                    yaxis_title="Project",
                )
                st.plotly_chart(fig_proj_cost, use_container_width=True)

            gc_ot, gc4 = st.columns(2)
            with gc_ot:
                df_ot_loc_month = (
                    df_chart.groupby(["Work Location", "Month"])[
                        "Parsed_Overtime"
                    ]
                    .sum()
                    .reset_index()
                )
                df_ot_loc_month["Text_Format"] = df_ot_loc_month[
                    "Parsed_Overtime"
                ].apply(format_rp_short)

                fig_ot_loc = px.bar(
                    df_ot_loc_month,
                    x="Work Location",
                    y="Parsed_Overtime",
                    color="Month",
                    barmode="group",
                    title=(
                        "Perbandingan Overtime Work Location (Compare per"
                        " Bulan)"
                    ),
                    text="Text_Format",
                )
                fig_ot_loc.update_traces(
                    textangle=0,
                    textposition="outside",
                    hovertemplate=(
                        "<b>Location:</b> %{x}<br><b>Overtime:</b> Rp"
                        " %{y:,.0f}<extra></extra>"
                    ),
                )
                fig_ot_loc.update_layout(
                    xaxis_title="Work Location",
                    yaxis_title="Total Overtime (Rp)",
                    xaxis_tickangle=-25,
                )
                st.plotly_chart(fig_ot_loc, use_container_width=True)

            with gc4:
                top_hc = (
                    df_chart.groupby("Project")["Name"]
                    .count()
                    .nlargest(10)
                    .reset_index(name="Headcount")
                )
                fig_hc = px.bar(
                    top_hc,
                    x="Headcount",
                    y="Project",
                    orientation="h",
                    title="Top 10 Project Berdasarkan Jumlah Headcount",
                    text_auto=True,
                    color="Headcount",
                    color_continuous_scale="Teal",
                )
                fig_hc.update_traces(textangle=0)
                fig_hc.update_layout(
                    yaxis={"categoryorder": "total ascending"},
                    xaxis_title="Jumlah Karyawan",
                    yaxis_title="Project",
                )
                st.plotly_chart(fig_hc, use_container_width=True)

            gc3, gc_emp = st.columns(2)
            with gc3:
                top_projects_list = (
                    df_chart.groupby("Project")["Parsed_Payment"]
                    .sum()
                    .nlargest(8)
                    .index
                )
                df_proj_month = (
                    df_chart[df_chart["Project"].isin(top_projects_list)]
                    .groupby(["Project", "Month"])["Parsed_Payment"]
                    .sum()
                    .reset_index()
                )
                df_proj_month["Text_Format"] = df_proj_month[
                    "Parsed_Payment"
                ].apply(format_rp_short)

                fig_proj_month = px.bar(
                    df_proj_month,
                    x="Project",
                    y="Parsed_Payment",
                    color="Month",
                    barmode="group",
                    title="Perbandingan Biaya Project (Compare per Bulan)",
                    text="Text_Format",
                )
                fig_proj_month.update_traces(
                    textangle=0,
                    textposition="outside",
                    hovertemplate=(
                        "<b>Project:</b> %{x}<br><b>Total Payment:</b> Rp"
                        " %{y:,.0f}<extra></extra>"
                    ),
                )
                fig_proj_month.update_layout(
                    xaxis_title="Project",
                    yaxis_title="Total Payment (Rp)",
                    xaxis_tickangle=-25,
                )
                st.plotly_chart(fig_proj_month, use_container_width=True)

            with gc_emp:
                # --- GRAFIK BAR PERBANDINGAN FDW VS TDW PER PROJECT ---
                if (
                    "Employment Status" in df_chart.columns
                    and "Project" in df_chart.columns
                ):
                    df_emp_chart = df_chart.copy()
                    df_emp_chart["Employment Status Clean"] = (
                        df_emp_chart["Employment Status"]
                        .astype(str)
                        .str.strip()
                        .str.upper()
                        .replace("", "BELUM DIISI")
                    )

                    # Ambil top 8 project terbesar untuk scannability
                    top_emp_projects = (
                        df_emp_chart.groupby("Project")["Parsed_Payment"]
                        .sum()
                        .nlargest(8)
                        .index
                    )

                    df_stat_proj = (
                        df_emp_chart[
                            df_emp_chart["Project"].isin(top_emp_projects)
                        ]
                        .groupby(["Project", "Employment Status Clean"])[
                            "Parsed_Payment"
                        ]
                        .sum()
                        .reset_index()
                    )
                    df_stat_proj["Text_Format"] = df_stat_proj[
                        "Parsed_Payment"
                    ].apply(format_rp_short)

                    fig_stat_proj = px.bar(
                        df_stat_proj,
                        x="Project",
                        y="Parsed_Payment",
                        color="Employment Status Clean",
                        barmode="group",
                        title=(
                            "Perbandingan Biaya (Total Payment) FDW vs TDW per"
                            " Project"
                        ),
                        text="Text_Format",
                        color_discrete_sequence=px.colors.qualitative.Bold,
                    )
                    fig_stat_proj.update_traces(
                        textangle=0,
                        textposition="outside",
                        textfont=dict(size=11),
                        cliponaxis=False,  # Properti ditaruh langsung di update_traces untuk mencegah error
                        hovertemplate=(
                            "<b>Project:</b> %{x}<br><b>Status:</b>"
                            " %{fullData.name}<br><b>Total Payment:</b> Rp"
                            " %{y:,.0f}<extra></extra>"
                        ),
                    )
                    fig_stat_proj.update_layout(
                        xaxis_title="Project",
                        yaxis_title="Total Payment (Rp)",
                        xaxis_tickangle=-25,
                        legend_title_text="Employment Status",
                        margin=dict(t=60, b=50),  # Memberikan margin atas yang cukup
                    )
                    st.plotly_chart(fig_stat_proj, use_container_width=True)

        with st.expander(
            "🔍 Cek Detail / Baris Angka Total Payment Amount", expanded=False
        ):
            st.caption(
                "Gunakan tabel ini untuk memastikan apakah ada baris yang"
                " bernilai 0 atau salah format."
            )
            debug_df = filtered_mc[[
                "Month",
                "Project",
                "Name",
                "Total Payment Amount",
            ]].copy()
            debug_df["Parsed Numeric"] = to_num(
                filtered_mc["Total Payment Amount"]
            )
            st.dataframe(debug_df, use_container_width=True)

        display_matrix = filtered_mc.copy()
        financial_cols = [
            "Basic Salary",
            "Meals & Transp",
            "Overtime",
            "Position",
            "Skill",
            "Other",
            "Shortage",
            "Deduction",
            "Total Salary",
            "BPJS",
            "Management Fee",
            "Total Manpower Cost",
            "Grand Total",
            "CJI",
            "KHQ Report",
            "PPN",
            "PPH23",
            "Total Payment Amount",
        ]

        for col in financial_cols:
            if col in display_matrix.columns:
                display_matrix[col] = display_matrix[col].apply(
                    lambda x: (
                        f"{int(round(to_num(pd.Series([x])).iloc[0])):,}".replace(
                            ",", "."
                        )
                        if str(x).strip() not in ["", "nan", "none", "-"]
                        else ""
                    )
                )

        st.subheader("📋 Data Manpower Cost Matrix")
        st.dataframe(display_matrix, use_container_width=True, height=450)

        st.download_button(
            label="📊 Download Data Manpower Cost (CSV)",
            data=filtered_mc.to_csv(index=False).encode("utf-8-sig"),
            file_name=(
                f"Manpower_Cost_Export_{date.today().strftime('%Y%m%d')}.csv"
            ),
            mime="text/csv",
        )


# ==============================================================================
# MODUL 4: AI HR ASSISTANT (PINTAR & BACA SELURUH COST CENTER / PROJECT)
# ==============================================================================
if menu_pilihan == "🤖 AI HR Assistant":

    st.title("🤖 AI HR Assistant")
    st.caption(
        "Asisten AI cerdas terintegrasi dengan Database Karyawan & Manpower Cost."
    )

    groq_key = st.secrets.get("GROQ_API_KEY", "")

    if not groq_key:
        st.error(
            "⚠️ API Key Groq belum dikonfigurasi di secrets.toml! Silakan"
            " tambahkan GROQ_API_KEY."
        )
        st.stop()

    client = Groq(api_key=groq_key)

    # --- HITUNG KONTEKS DATA REALTIME UNTUK DIBERIKAN KE AI ---
    df_emp = st.session_state.get("employees", pd.DataFrame())

    # Filter Karyawan Aktif
    if not df_emp.empty and "Status" in df_emp.columns:
        df_aktif = df_emp[df_emp["Status"] == "Aktif"].copy()
        total_emp = len(df_emp)
        aktif_emp = len(df_aktif)
        resign_emp = len(df_emp[df_emp["Status"] == "Resign"])
    else:
        df_aktif = df_emp.copy() if not df_emp.empty else pd.DataFrame()
        total_emp = len(df_emp)
        aktif_emp, resign_emp = total_emp, 0

    # 1. Ringkasan per Site Total
    site_summary = ""
    if not df_aktif.empty and "Site" in df_aktif.columns:
        site_counts = (
            df_aktif["Site"]
            .astype(str)
            .str.strip()
            .str.upper()
            .value_counts()
            .to_dict()
        )
        site_summary = ", ".join([
            f"{k}: {v} orang"
            for k, v in site_counts.items()
            if k and k != "NAN"
        ])

    # 2. GROUPING RINCI: SITE & POSISI
    site_posisi_summary = ""
    if (
        not df_aktif.empty
        and "Site" in df_aktif.columns
        and "Posisi" in df_aktif.columns
    ):
        df_grouped = (
            df_aktif.groupby(["Site", "Posisi"])["ID"]
            .count()
            .reset_index(name="Jumlah")
        )
        grouped_items = []
        for _, row in df_grouped.iterrows():
            site_val = str(row["Site"]).strip().upper()
            pos_val = str(row["Posisi"]).strip()
            cnt_val = row["Jumlah"]
            if site_val and site_val != "NAN" and pos_val and pos_val != "NAN":
                grouped_items.append(f"[{site_val} - {pos_val}: {cnt_val} orang]")

        site_posisi_summary = ", ".join(grouped_items)

    # 3. Ringkasan LENGKAP Seluruh Cost Center dari Master Karyawan (Tanpa dipotong)
    cc_summary = ""
    if not df_aktif.empty and "Cost Center" in df_aktif.columns:
        df_cc_clean = (
            df_aktif["Cost Center"]
            .astype(str)
            .str.strip()
            .str.upper()
            .replace({"FKS": "FKS", "": "BELUM DIISI"})
        )
        cc_counts = df_cc_clean.value_counts().to_dict()
        cc_summary = ", ".join([
            f"{k}: {v} orang" for k, v in cc_counts.items() if k and k != "NAN"
        ])

    # 4. Ringkasan Data Project dari Tab Manpower Cost (jika terisi)
    mp_proj_summary = ""
    df_mc_session = st.session_state.get("df_manpower_cost", pd.DataFrame())
    if not df_mc_session.empty and "Project" in df_mc_session.columns:
        proj_counts = (
            df_mc_session["Project"]
            .astype(str)
            .str.strip()
            .str.upper()
            .value_counts()
            .to_dict()
        )
        mp_proj_summary = ", ".join([
            f"{k}: {v} orang" for k, v in proj_counts.items() if k and k != "NAN"
        ])

    # --- SYSTEM PROMPT LENGKAP BACA SELURUH COST CENTER & PROJECT ---
    system_prompt_context = f"""
    Anda adalah Asisten AI HR internal perusahaan yang cerdas, presisi, dan ramah.
    Anda memiliki akses langsung ke data realtime database berikut:

    📊 RINGKASAN UMUM:
    - Total Record Karyawan: {total_emp} orang
    - Karyawan Aktif: {aktif_emp} orang
    - Karyawan Resign: {resign_emp} orang
    - Total Karyawan per Site/Lokasi: {site_summary if site_summary else 'Belum ada data'}

    🔥 RINCIAN POSISI PER SITE (LOKASI KERJA):
    {site_posisi_summary if site_posisi_summary else 'Belum ada data detail'}

    💳 SELURUH COST CENTER / PROJECT (MASTER KARYAWAN):
    {cc_summary if cc_summary else 'Belum ada data'}

    📋 DATA PROJECT (TAB MANPOWER COST):
    {mp_proj_summary if mp_proj_summary else 'Belum ada data Manpower Cost'}

    PETUNJUK BALASAN:
    1. Jika pengguna menanyakan jumlah orang pada Cost Center atau Project tertentu (misalnya FKS, VinFast, CJ Food, dsb.), cocokkan dengan daftar "SELURUH COST CENTER / PROJECT" di atas dan sebutkan jumlahnya secara akurat.
    2. Jika pengguna menanyakan posisi spesifik di lokasi tertentu (misal: "posisi admin di JDC"), gunakan data "RINCIAN POSISI PER SITE".
    3. Jawab selalu dengan bahasa Indonesia yang sopan, ramah, dan profesional.
    """

    # Inisialisasi Chat History
    if "chat_messages" not in st.session_state:
        st.session_state.chat_messages = [
            {
                "role": "assistant",
                "content": (
                    f"Halo! Saya AI HR Assistant. Saya telah membaca seluruh"
                    f" database ({aktif_emp} Karyawan Aktif). Silakan tanyakan"
                    f" jumlah karyawan berdasarkan Cost Center, Project, maupun Lokasi Site!"
                ),
            }
        ]

    # Tampilkan riwayat chat
    for message in st.session_state.chat_messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])

    # Input User
    if prompt := st.chat_input(
        "Tanyakan sesuatu (misal: 'berapa orang yang pegang project FKS?')..."
    ):
        st.session_state.chat_messages.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.markdown(prompt)

        with st.chat_message("assistant"):
            with st.spinner("Mencari data Cost Center / Project..."):
                try:
                    api_messages = [
                        {"role": "system", "content": system_prompt_context}
                    ] + [
                        {"role": m["role"], "content": m["content"]}
                        for m in st.session_state.chat_messages
                    ]

                    completion = client.chat.completions.create(
                        model="llama-3.3-70b-versatile",
                        messages=api_messages,
                        temperature=0.2,
                        max_tokens=1024,
                    )

                    response_text = completion.choices[0].message.content
                    st.markdown(response_text)

                    st.session_state.chat_messages.append(
                        {"role": "assistant", "content": response_text}
                    )

                except Exception as e:
                    st.error(f"❌ Terjadi kesalahan pada AI: {e}")

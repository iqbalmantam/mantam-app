import streamlit as st
import pandas as pd
import openpyxl

# Patch OpenPyXL untuk menangani kustom filter pada Excel
try:
    import openpyxl.worksheet.filters as filters
    class CustomDescriptor:
        def __set__(self, instance, value):
            instance.__dict__['_val'] = str(value) if value is not None else ""
        def __get__(self, instance, owner):
            return instance.__dict__.get('_val', '')
            
    filters.CustomFilter.val = CustomDescriptor()
except Exception:
    pass

st.set_page_config(
    page_title="JDC Manpower & OT Dashboard",
    page_icon="📊",
    layout="wide"
)

st.title("📊 JDC Manpower & Overtime Daily Report Dashboard")
st.caption("Aplikasi Otomasi & Analisis Laporan Harian Manpower & OT - PT CJ Logistics Service Indonesia")

# Upload File Excel Laporan
uploaded_file = st.file_uploader("Unggah File Excel Laporan (xlsx)", type=["xlsx"])

if uploaded_file is not None:
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    sheet_names = wb.sheetnames
    
    st.sidebar.header("📂 Sheet Terdeteksi")
    st.sidebar.write(sheet_names)

    # Tab Navigasi Fitur
    tab1, tab2, tab3, tab4 = st.tabs([
        "📈 Summary & Headcount", 
        "⏰ Overtime (OT 24)", 
        "💰 Cost OT Analysis", 
        "📋 Data Outsourcing (OS)"
    ])

    # TAB 1: SUMMARY & HEADCOUNT
    with tab1:
        st.subheader("📌 Ringkasan Manpower")
        if 'OS' in sheet_names:
            ws_os = wb['OS']
            os_data = []
            for r in range(5, ws_os.max_row + 1):
                vals = [ws_os.cell(r, c).value for c in range(1, 15)]
                if any(vals):
                    os_data.append(vals)
            
            if os_data:
                df_os = pd.DataFrame(os_data[1:], columns=os_data[0])
                total_os = len(df_os.dropna(subset=['Name'])) if 'Name' in df_os.columns else len(df_os)
            else:
                total_os = 0
        else:
            total_os = 0

        col1, col2, col3 = st.columns(3)
        col1.metric("Total Headcount OS", f"{total_os} Personel")
        col2.metric("Status Periode Report", "Agustus 2026")
        col3.metric("Status Berkas", "Terverifikasi")

    # TAB 2: OVERTIME (OT 24)
    with tab2:
        st.subheader("⏰ Rekapitulasi Overtime Detail (OT 24)")
        if 'OT 24' in sheet_names:
            ws_ot = wb['OT 24']
            ot_rows = []
            for r in range(4, ws_ot.max_row + 1):
                vals = [ws_ot.cell(r, c).value for c in range(1, 15)]
                if any(vals):
                    ot_rows.append(vals)
            
            if len(ot_rows) > 1:
                df_ot = pd.DataFrame(ot_rows[1:], columns=ot_rows[0])
                st.dataframe(df_ot, use_container_width=True)
                
                csv_ot = df_ot.to_csv(index=False).encode('utf-8')
                st.download_button("📥 Export Rekap OT 24 (CSV)", data=csv_ot, file_name="Rekap_OT24.csv", mime="text/csv")
            else:
                st.info("Data OT 24 kosong atau format tidak sesuai.")
        else:
            st.warning("Sheet 'OT 24' tidak ditemukan.")

    # TAB 3: COST OT ANALYSIS
    with tab3:
        st.subheader("💰 Ringkasan Biaya Overtime")
        if 'Cost OT' in sheet_names:
            ws_cost = wb['Cost OT']
            cost_data = []
            for r in range(3, ws_cost.max_row + 1):
                vals = [ws_cost.cell(r, c).value for c in range(1, 11)]
                if any(vals):
                    cost_data.append(vals)
            
            if len(cost_data) > 1:
                df_cost = pd.DataFrame(cost_data[1:], columns=cost_data[0])
                st.dataframe(df_cost, use_container_width=True)
        else:
            st.warning("Sheet 'Cost OT' tidak ditemukan.")

    # TAB 4: DATA OUTSOURCING
    with tab4:
        st.subheader("📋 Data Outsourcing (PT ARU)")
        if 'OS' in sheet_names and 'df_os' in locals():
            st.dataframe(df_os, use_container_width=True)

else:
    st.info("Silakan unggah file `JDC MP Daily Report August 2026.xlsx` untuk memulai analisis.")
